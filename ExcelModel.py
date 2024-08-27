import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def crearExcel():
    
    if comprobarSiExiste():
        print("El archivo de excel ya existe, por favor, borralo para crearlo desde cero")
    else:
        # Crear un nuevo libro de trabajo
        workbook = Workbook()
        # Seleccionar la hoja activa
        sheet = workbook.active
        
        # Especificar los encabezados
        encabezados = ["id", "nombre", "mail", "asunto", "cuerpo_mensaje", "path_adjunto", "estado_mail", "fecha_envio"]
        
        # Asignar los encabezados en la fila 1
        sheet.append(encabezados)
        
        # Guardar el archivo Excel en el directorio actual con el nombre "datos_mensajes.xlsx"
        workbook.save("datos_mensajes.xlsx")


## simplemente comprueba si el archivo existe, para que devuelva un mensaje de error
## cuando sea necesario. 
def comprobarSiExiste():
    # Verificar si el archivo "datos_mensajes.xlsx" existe en el directorio actual
    return os.path.isfile("datos_mensajes.xlsx")



### según los datos de la excel, busca nombre en el archivo pdf y se lo asigna para adjuntarlo. 
def escribirPathsAdjuntos():
    # Cargar el archivo Excel
    workbook = load_workbook("datos_mensajes.xlsx")
    # Seleccionar la hoja activa
    sheet = workbook.active
    
    # Crear una lista para almacenar los valores de la columna B
    nombres_columna_b = []

    # Iterar sobre las filas a partir de la segunda (para omitir los encabezados)
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):  # Columna B es la segunda (min_col=2, max_col=2)
        nombre = row[0].value  # Obtener el valor de la columna B (solo hay una celda en la columna)
        if nombre:
            nombres_columna_b.append(nombre)  # Almacenar en la lista
        
    # Iterar de nuevo para escribir en la columna F
    for idx, nombre in enumerate(nombres_columna_b, start=2):  # Comenzar en la fila 2
        # Buscar el archivo PDF cuyo nombre contenga el valor de la columna B
        pdf_path = None
        for root, dirs, files in os.walk("pdfs"):
            for file in files:
                if file.endswith(".pdf") and nombre in file:
                    pdf_path_full = os.path.join(root, file)
                    pdf_path = os.path.join(file)
                    break
            if pdf_path:
                break
        
        # Escribir el path en la columna F de la fila correspondiente
        sheet.cell(row=idx, column=6).value = pdf_path
    
    # Guardar los cambios en el archivo Excel
    workbook.save("datos_mensajes.xlsx")

## devuelve la lista de envio para poder trabajar con ella. Todas las columnas. 
def getListaEnvio():
    # Cargar el archivo Excel
    workbook = load_workbook("datos_mensajes.xlsx")
    # Seleccionar la hoja activa
    sheet = workbook.active
    
    # Inicializar la lista que contendrá los datos
    lista_envio = []

    # Iterar sobre las filas a partir de la segunda (para omitir los encabezados)
    for row in sheet.iter_rows(min_row=2, values_only=True):  # values_only=True retorna solo los valores, no los objetos de celda
        lista_envio.append(list(row))  # Convertir la fila a una lista y agregarla a lista_envio
    
    return lista_envio

### al recibir un listado de IDs, los colocará como enviados y les pondrá la fecha de envio. 
def ponerComoEnviados(lista_ids):
    # Cargar el archivo Excel
    workbook = load_workbook("datos_mensajes.xlsx")
    # Seleccionar la hoja activa
    sheet = workbook.active

    # Obtener la fecha de hoy
    fecha_hoy = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    
    # Iterar sobre las filas a partir de la segunda (para omitir los encabezados)
    for row in sheet.iter_rows(min_row=2):
        # Obtener el ID de la fila actual (columna A)
        id_fila = row[0].value  # La columna A es la primera (índice 0)
        
        # Si el ID está en la lista de IDs proporcionada, actualizar el estado
        if id_fila in lista_ids:
            row[6].value = "Enviado"  # La columna "estado_mail" es la séptima (índice 6)
            row[7].value = fecha_hoy  # La columna "fecha_envio" es la octava (índice 7)
    
    # Guardar los cambios en el archivo Excel
    workbook.save("datos_mensajes.xlsx")

def isExcelAbierto():
    try:
        # Intentar abrir el archivo en modo de escritura exclusiva
        with open("datos_mensajes.xlsx"), 'r+':
            pass
    except IOError:
        # Si ocurre un error, significa que el archivo está abierto
        return True
    
def cerrarExcel():
    if not isExcelAbierto():
        workbook = load_workbook("datos_mensajes.xlsx")
        workbook.save("datos_mensajes.xlsx")
    else:
        print("El archivo de Excel no está abierto")